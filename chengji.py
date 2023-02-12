import openpyxl
from openpyxl.styles import PatternFill
import xlrd
print("使用须知\n1.必须剔除不评选奖学金的学生（包括转专业和复学的同学，因为他们在本班的必修课中无成绩）\n2.剔除上述学生的必修和专业选修\n3.为必修成绩中为空的单元格设置数值为0\n4.计算后的成绩需要你手动排序")
print("**********************")
print("\t\t\t作者：杜志力")
print("\t\t\t版本：1.0")
print("\t\t\t日期：2023.2.12")
print("开源：https://github.com/2962283762/cdtu-jiangxuejinjisuan.git")
fille=PatternFill('solid', fgColor='ff0000')
gongyifenfilename = input('请输入公益分文件名：')
caoxingfenfilename = input('请输入操行分文件名：')
chengjifilename = input('请输入成绩文件名：')
def getgongyifen(gongyifenfilename):
    book = xlrd.open_workbook(gongyifenfilename)
    sheet = book.sheet_by_index(0)
    clo57 = sheet.col_values(sheet.merged_cells[1][3])
    nrows = sheet.nrows
    gongyifen = []
    tep = []
    merged=sheet.merged_cells
    for i in range(0,len(merged)):
        if(merged[i][3]==57 and merged[i][2]==56):
            tep.append(merged[i]) 
        
    for i in range(0,len(tep)):
        gongyifen.append(sheet.cell_value(tep[i][0],57))
    return gongyifen
def getcaoxingfen(caoxingfenfilename):
    book = xlrd.open_workbook(caoxingfenfilename)
    sheet = book.sheets()[0]
    nrows = sheet.nrows
    clo20 = sheet.col_values(19)
    caoxingfen = []
    for i in range(3,nrows):
        caoxingfen.append(clo20[i])
    return caoxingfen
book = openpyxl.load_workbook(chengjifilename)
sheet = book.active
clos_num = sheet.max_column
row_num = sheet.max_row
for i in range(clos_num,3,-1):
    value = sheet.cell(row=3,column=i).value
    if(value.find('必修')==-1 and value.find('专业选修')==-1):
        # print(value + "\t" + str(i))
        sheet.delete_cols(i)
PE = []
sheet.delete_cols(3)
sheet.delete_rows(row_num-1,row_num)
clos_num = sheet.max_column
row_num = sheet.max_row
credit_list = []
credit_sum=0
score_sum=0
for i in range(3,clos_num+1):
    credit=sheet.cell(row=3,column=i).value
    if(credit.find('体育')==-1):
        credit=credit.split('，')
        credit = credit[3].replace('学分','')
        credit = credit.replace(' ','')
        credit= float(credit)
        credit_sum+=credit
        credit_list.append(credit)
    

sheet.cell(row=3,column=clos_num+1).value='成绩总分'
sheet.cell(row=3,column=clos_num+2).value='公益分'
sheet.cell(row=3,column=clos_num+3).value='操行分'
sheet.cell(row=3,column=clos_num+4).value='奖学金分数'
print(credit_list)
for i in range(4,row_num+1):
    score_sum=0
    score_list=[]
    for j in range(3,clos_num+1):
        if(sheet.cell(row=3,column=j).value.find('体育')==-1):
            score_list.append(sheet.cell(row=i,column=j).value)
            if(float(sheet.cell(row=i,column=j).value)<60):
                for s in range(1,clos_num+5):
                    sheet.cell(row=i,column=s).fill=fille
        else:
            PE.append(float(sheet.cell(row=i,column=j).value))
            if(float(sheet.cell(row=i,column=j).value)<60):
               for s in range(1,clos_num+5):
                    sheet.cell(row=i,column=s).fill=fille
    for k in range(len(score_list)):
        score_sum+=float(score_list[k])*credit_list[k]
    sheet.cell(row=i,column=clos_num+1).value=str(score_sum/credit_sum)
gongyifen=getgongyifen(gongyifenfilename)
for i in range(4,row_num+1):
    print(gongyifen)
    if(gongyifen[i-4]>=300 and gongyifen[i-4]<1000):
        sheet.cell(row=i,column=clos_num+2).value = (60+(gongyifen[i-4]-300)*40/700)*0.09
    elif(gongyifen[i-4]>=1000):
        sheet.cell(row=i,column=clos_num+2).value = 9
    else:
        sheet.cell(row=i,column=clos_num+2).value = gongyifen[i-4]*0.09*0.2
        for s in range(1,clos_num+5):
                    sheet.cell(row=i,column=s).fill=fille
        
    
caoxingfen=getcaoxingfen(caoxingfenfilename)

for i in range(4,row_num+1):
    sheet.cell(row=i,column=clos_num+3).value = caoxingfen[i-4]
    if(caoxingfen[i-4]<85):
        for s in range(1,clos_num+5):
                    sheet.cell(row=i,column=s).fill=fille
print(caoxingfen)
choice = input('请选择年级：\n1.大一\n2.大二\n3.大三\n4.大四\n')
if(choice=='1' or choice=='2'):
    for j in range(4,row_num+1):
        
        sheet.cell(row=j,column=clos_num+4).value = float(sheet.cell(row=j,column=clos_num+1).value)*0.7+float(sheet.cell(row=j,column=clos_num+2).value)+float(sheet.cell(row=j,column=clos_num+3).value)*0.11+PE[j-4]*0.1
elif(choice=='3'):
    for j in range(4,row_num+1):
        sheet.cell(row=j,column=clos_num+4).value = float(sheet.cell(row=j,column=clos_num+1).value)*0.8+float(sheet.cell(row=j,column=clos_num+2).value)+float(sheet.cell(row=j,column=clos_num+3).value)*0.11

else:
    for j in range(4,row_num+1):
        sheet.cell(row=j,column=clos_num+4).value = float(sheet.cell(row=j,column=clos_num+1).value)*0.89+float(sheet.cell(row=j,column=clos_num+3).value)*0.11

    
book.save('奖学金表.xlsx')